import io
import json
import mimetypes
import smtplib
import ssl
from email.message import EmailMessage
from urllib import error, request

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Font


st.set_page_config(page_title="화장품 영업 데이터 분석기", layout="wide")


REQUIRED_HINTS = {
    "date": ["날짜", "일자", "date"],
    "product": ["제품명", "상품명", "품목명", "product"],
    "sales": ["매출액", "매출", "sales", "revenue"],
    "margin": ["영업이익률", "이익률", "마진율", "margin"],
}


def find_column(columns, hints):
    normalized = {str(col).strip().lower(): col for col in columns}
    for hint in hints:
        key = hint.strip().lower()
        for normalized_name, original_name in normalized.items():
            if key == normalized_name or key in normalized_name:
                return original_name
    return None


def validate_columns(df):
    matched = {key: find_column(df.columns, hints) for key, hints in REQUIRED_HINTS.items()}
    missing = [key for key, value in matched.items() if value is None]
    return matched, missing


def clean_numeric(series):
    cleaned = (
        series.astype(str)
        .str.replace(",", "", regex=False)
        .str.replace("%", "", regex=False)
        .str.replace("원", "", regex=False)
        .str.strip()
    )
    return pd.to_numeric(cleaned, errors="coerce")


def process_uploaded_excel(uploaded_file):
    source_bytes = uploaded_file.getvalue()
    excel_file = pd.ExcelFile(io.BytesIO(source_bytes))
    source_df = pd.read_excel(excel_file, sheet_name=0)
    matched, missing = validate_columns(source_df)
    if missing:
        labels = {
            "date": "날짜",
            "product": "제품명",
            "sales": "매출액",
            "margin": "영업이익률",
        }
        missing_text = ", ".join(labels[item] for item in missing)
        raise ValueError(f"필수 컬럼을 찾지 못했습니다: {missing_text}")

    work_df = source_df.copy()
    sales_col_name = str(matched["sales"])
    sales_unit = "만원" if "만원" in sales_col_name else "원"
    work_df["정규화_날짜"] = pd.to_datetime(work_df[matched["date"]], errors="coerce")
    work_df["정규화_제품명"] = work_df[matched["product"]].astype(str).str.strip()
    work_df["정규화_매출액"] = clean_numeric(work_df[matched["sales"]])
    work_df["정규화_영업이익률"] = clean_numeric(work_df[matched["margin"]])

    valid_df = work_df.dropna(subset=["정규화_날짜", "정규화_매출액"]).copy()
    valid_df["날짜"] = valid_df["정규화_날짜"].dt.date

    sales_by_date = (
        valid_df.groupby("날짜", as_index=False)["정규화_매출액"]
        .sum()
        .rename(columns={"정규화_매출액": "매출액"})
        .sort_values("날짜")
    )
    sales_by_product = (
        valid_df.groupby("정규화_제품명", as_index=False)["정규화_매출액"]
        .sum()
        .rename(columns={"정규화_제품명": "제품명", "정규화_매출액": "매출액"})
        .sort_values("매출액", ascending=False)
    )
    margin_by_product = (
        work_df.dropna(subset=["정규화_제품명", "정규화_영업이익률"])
        .groupby("정규화_제품명", as_index=False)["정규화_영업이익률"]
        .mean()
        .rename(columns={"정규화_제품명": "제품명", "정규화_영업이익률": "평균 영업이익률"})
        .sort_values("평균 영업이익률", ascending=False)
    )
    avg_sales_by_date = (
        valid_df.groupby("날짜", as_index=False)["정규화_매출액"]
        .mean()
        .rename(columns={"정규화_매출액": "평균 매출액"})
        .sort_values("날짜")
    )

    workbook = load_workbook(io.BytesIO(source_bytes))
    for sheet_name in ["날짜별_매출액", "제품별_매출액", "분석_차트"]:
        if sheet_name in workbook.sheetnames:
            del workbook[sheet_name]

    ws_date = workbook.create_sheet("날짜별_매출액")
    ws_product = workbook.create_sheet("제품별_매출액")
    ws_chart = workbook.create_sheet("분석_차트")

    append_dataframe(ws_date, sales_by_date, "날짜별 매출액")
    append_dataframe(ws_product, sales_by_product, "제품명별 매출액")

    ws_chart["A1"] = "제품명별 평균 영업이익률"
    ws_chart["A1"].font = Font(bold=True, size=12)
    append_table(ws_chart, margin_by_product, start_row=2)

    avg_sales_start = max(len(margin_by_product) + 5, 12)
    ws_chart.cell(row=avg_sales_start, column=1, value="날짜별 평균 매출액").font = Font(bold=True, size=12)
    append_table(ws_chart, avg_sales_by_date, start_row=avg_sales_start + 1)

    add_bar_chart(ws_chart, len(margin_by_product))
    add_line_chart(ws_chart, avg_sales_start + 1, len(avg_sales_by_date))

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    summary = {
        "원본 행 수": int(len(source_df)),
        "유효 행 수": int(len(valid_df)),
        "총 매출액": float(valid_df["정규화_매출액"].sum()),
        "평균 영업이익률": float(work_df["정규화_영업이익률"].dropna().mean() or 0),
        "제품 수": int(valid_df["정규화_제품명"].nunique()),
        "기간": f"{sales_by_date['날짜'].min()} ~ {sales_by_date['날짜'].max()}" if not sales_by_date.empty else "-",
        "매출 단위": sales_unit,
    }

    chat_context = {
        "summary": summary,
        "sales_by_date": sales_by_date.assign(날짜=sales_by_date["날짜"].astype(str)).to_dict(orient="records"),
        "sales_by_product": sales_by_product.head(20).to_dict(orient="records"),
        "margin_by_product": margin_by_product.head(20).to_dict(orient="records"),
        "avg_sales_by_date": avg_sales_by_date.assign(날짜=avg_sales_by_date["날짜"].astype(str)).to_dict(orient="records"),
    }

    return {
        "processed_bytes": output.getvalue(),
        "sales_by_date": sales_by_date,
        "sales_by_product": sales_by_product,
        "margin_by_product": margin_by_product,
        "avg_sales_by_date": avg_sales_by_date,
        "summary": summary,
        "chat_context": chat_context,
    }


def append_dataframe(ws, df, title):
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=12)
    append_table(ws, df, start_row=2)


def append_table(ws, df, start_row):
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=start_row, column=col_idx, value=str(col_name))
        cell.font = Font(bold=True)

    for row_offset, row_values in enumerate(df.itertuples(index=False), start=1):
        for col_idx, value in enumerate(row_values, start=1):
            ws.cell(row=start_row + row_offset, column=col_idx, value=value)


def add_bar_chart(ws, data_len):
    if data_len == 0:
        return
    chart = BarChart()
    chart.title = "제품명별 평균 영업이익률"
    chart.y_axis.title = "영업이익률"
    chart.x_axis.title = "제품명"
    data = Reference(ws, min_col=2, min_row=2, max_row=data_len + 2)
    categories = Reference(ws, min_col=1, min_row=3, max_row=data_len + 2)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    chart.height = 8
    chart.width = 14
    ws.add_chart(chart, "G5")


def add_line_chart(ws, table_start_row, data_len):
    if data_len == 0:
        return
    chart = LineChart()
    chart.title = "날짜별 평균 매출액"
    chart.y_axis.title = "평균 매출액"
    chart.x_axis.title = "날짜"
    data = Reference(ws, min_col=2, min_row=table_start_row, max_row=table_start_row + data_len)
    categories = Reference(ws, min_col=1, min_row=table_start_row + 1, max_row=table_start_row + data_len)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    chart.height = 8
    chart.width = 14
    ws.add_chart(chart, "G10")


def send_email_with_attachment(
    smtp_host,
    smtp_port,
    sender_email,
    sender_password,
    recipient_email,
    subject,
    body,
    attachment_name,
    attachment_bytes,
):
    message = EmailMessage()
    message["Subject"] = subject
    message["From"] = sender_email
    message["To"] = recipient_email
    message.set_content(body)

    mime_type, _ = mimetypes.guess_type(attachment_name)
    maintype, subtype = (mime_type or "application/octet-stream").split("/", 1)
    message.add_attachment(attachment_bytes, maintype=maintype, subtype=subtype, filename=attachment_name)

    context = ssl.create_default_context()
    with smtplib.SMTP_SSL(smtp_host, smtp_port, context=context) as server:
        server.login(sender_email, sender_password)
        server.send_message(message)


def build_local_chat_answer(question, context):
    summary = context["summary"]
    product_rows = context["sales_by_product"]
    date_rows = context["sales_by_date"]
    sales_unit = summary.get("매출 단위", "원")

    if "총매출" in question or "전체 매출" in question:
        return f"총 매출액은 {summary['총 매출액']:,.0f}{sales_unit}입니다."
    if "가장 많이 팔린 제품" in question or "매출 1위" in question:
        if product_rows:
            top = product_rows[0]
            return f"매출 1위 제품은 {top['제품명']}이며 매출액은 {top['매출액']:,.0f}{sales_unit}입니다."
    if "가장 매출이 높은 날짜" in question or "매출이 가장 높은 날짜" in question:
        if date_rows:
            top = max(date_rows, key=lambda row: row["매출액"])
            return f"매출이 가장 높은 날짜는 {top['날짜']}이고 매출액은 {top['매출액']:,.0f}{sales_unit}입니다."

    lines = [
        "OpenAI API 키가 없어서 로컬 요약 기반으로 답변드리고 있습니다.",
        f"- 분석 기간: {summary['기간']}",
        f"- 총 매출액: {summary['총 매출액']:,.0f}{sales_unit}",
        f"- 제품 수: {summary['제품 수']}개",
    ]
    if product_rows:
        top3 = ", ".join(f"{row['제품명']}({row['매출액']:,.0f}{sales_unit})" for row in product_rows[:3])
        lines.append(f"- 상위 제품: {top3}")
    return "\n".join(lines)


def ask_openai(api_key, question, context):
    payload = {
        "model": "gpt-4.1-mini",
        "input": [
            {
                "role": "system",
                "content": [
                    {
                        "type": "input_text",
                        "text": "당신은 엑셀 영업 데이터 분석 도우미입니다. 주어진 JSON 데이터만 근거로 한국어로 간결하게 답변하세요.",
                    }
                ],
            },
            {
                "role": "user",
                "content": [
                    {
                        "type": "input_text",
                        "text": f"데이터:\n{json.dumps(context, ensure_ascii=False)}\n\n질문:\n{question}",
                    }
                ],
            },
        ],
    }
    req = request.Request(
        "https://api.openai.com/v1/responses",
        data=json.dumps(payload).encode("utf-8"),
        headers={
            "Content-Type": "application/json",
            "Authorization": f"Bearer {api_key}",
        },
        method="POST",
    )
    try:
        with request.urlopen(req, timeout=60) as response:
            result = json.loads(response.read().decode("utf-8"))
            return result["output"][0]["content"][0]["text"]
    except error.HTTPError as exc:
        detail = exc.read().decode("utf-8", errors="ignore")
        raise RuntimeError(f"OpenAI API 호출 실패: {detail}") from exc
    except Exception as exc:
        raise RuntimeError(f"OpenAI API 호출 실패: {exc}") from exc


def render_overview(result):
    sales_unit = result["summary"].get("매출 단위", "원")
    col1, col2, col3 = st.columns(3)
    col1.metric("총 매출액", f"{result['summary']['총 매출액']:,.0f}{sales_unit}")
    col2.metric("제품 수", f"{result['summary']['제품 수']}개")
    col3.metric("분석 기간", result["summary"]["기간"])
    st.caption(f"이 파일의 매출 단위는 `{sales_unit}` 기준입니다.")

    left, right = st.columns(2)
    with left:
        st.subheader("날짜별 매출액")
        st.dataframe(result["sales_by_date"], use_container_width=True)
    with right:
        st.subheader("제품명별 매출액")
        st.dataframe(result["sales_by_product"], use_container_width=True)


def render_download_and_email(result, original_name):
    processed_name = f"가공완료_{original_name}"
    st.download_button(
        "가공된 엑셀 다운로드",
        data=result["processed_bytes"],
        file_name=processed_name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )

    st.markdown("---")
    st.subheader("엑셀 메일 발송")
    with st.form("email_form"):
        col1, col2 = st.columns(2)
        smtp_host = col1.text_input("SMTP 서버", value="smtp.gmail.com")
        smtp_port = col2.number_input("SMTP 포트", min_value=1, max_value=65535, value=465)
        sender_email = st.text_input("보내는 사람 이메일")
        sender_password = st.text_input("앱 비밀번호 또는 SMTP 비밀번호", type="password")
        recipient_email = st.text_input("받는 사람 이메일")
        subject = st.text_input("메일 제목", value="화장품 영업 데이터 가공 파일")
        body = st.text_area("메일 본문", value="가공된 엑셀 파일을 첨부드립니다.")
        submitted = st.form_submit_button("메일 발송")

    if submitted:
        required_values = [smtp_host, sender_email, sender_password, recipient_email, subject]
        if not all(required_values):
            st.error("메일 발송에 필요한 정보를 모두 입력해주세요.")
        else:
            try:
                send_email_with_attachment(
                    smtp_host=smtp_host,
                    smtp_port=int(smtp_port),
                    sender_email=sender_email,
                    sender_password=sender_password,
                    recipient_email=recipient_email,
                    subject=subject,
                    body=body,
                    attachment_name=processed_name,
                    attachment_bytes=result["processed_bytes"],
                )
                st.success("메일을 발송했습니다.")
            except Exception as exc:
                st.error(f"메일 발송 중 오류가 발생했습니다: {exc}")


def render_chatbot(result):
    st.subheader("엑셀 기반 챗봇")
    st.caption("OpenAI API 키를 입력하면 더 자연스러운 답변을 받을 수 있고, 비워두면 로컬 요약 기반으로 답변합니다.")

    api_key = st.text_input("OpenAI API Key (선택)", type="password")
    if "chat_history" not in st.session_state:
        st.session_state.chat_history = []

    for item in st.session_state.chat_history:
        with st.chat_message(item["role"]):
            st.markdown(item["content"])

    question = st.chat_input("예: 매출 1위 제품은 뭐야?")
    if question:
        st.session_state.chat_history.append({"role": "user", "content": question})
        with st.chat_message("user"):
            st.markdown(question)

        with st.chat_message("assistant"):
            with st.spinner("답변 생성 중..."):
                try:
                    if api_key:
                        answer = ask_openai(api_key, question, result["chat_context"])
                    else:
                        answer = build_local_chat_answer(question, result["chat_context"])
                except Exception as exc:
                    answer = f"답변 생성 중 오류가 발생했습니다: {exc}"
                st.markdown(answer)
        st.session_state.chat_history.append({"role": "assistant", "content": answer})


def main():
    st.title("화장품 영업 데이터 자동 분석")
    st.write("엑셀 업로드 한 번으로 집계 시트 생성, 차트 삽입, 다운로드, 메일 발송, 챗봇까지 처리합니다.")

    uploaded_file = st.file_uploader("`화장품_영업_데이터` 엑셀 업로드", type=["xlsx", "xlsm", "xltx", "xltm"])
    if not uploaded_file:
        st.info("먼저 엑셀 파일을 업로드해주세요.")
        with st.expander("필수 컬럼 안내"):
            st.write("다음 의미의 컬럼이 필요합니다: 날짜, 제품명, 매출액, 영업이익률")
            st.write("예시 컬럼명: `날짜`, `제품명`, `매출액`, `영업이익률`")
        return

    try:
        result = process_uploaded_excel(uploaded_file)
    except Exception as exc:
        st.error(f"파일 처리 중 오류가 발생했습니다: {exc}")
        return

    tabs = st.tabs(["분석 결과", "다운로드/메일", "챗봇"])
    with tabs[0]:
        render_overview(result)
    with tabs[1]:
        render_download_and_email(result, uploaded_file.name)
    with tabs[2]:
        render_chatbot(result)


if __name__ == "__main__":
    main()
